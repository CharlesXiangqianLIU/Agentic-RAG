import io

import openpyxl

from frontend.export import chunks_to_excel
from frontend.export_md import conversation_to_markdown


def test_excel_export_returns_bytes():
    chunks = [
        {
            "text": "Entry 3 | Pd(OAc)2 | 80 °C | 87%",
            "attribution": "[Source: PRJ-031.docx | Page 12 | Section: Table 1]",
            "payload": {"source_file": "PRJ-031.docx", "page_number": 12, "section": "Table 1"},
        }
    ]
    result = chunks_to_excel("What was the yield?", chunks)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_excel_export_empty_chunks():
    result = chunks_to_excel("A question", [])
    assert isinstance(result, bytes)


def test_excel_has_question_in_first_row():
    chunks = [{"text": "87%", "attribution": "", "payload": {"source_file": "a.docx", "page_number": 1, "section": "S1"}}]
    data = chunks_to_excel("Test question?", chunks)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Question"
    assert ws.cell(row=1, column=2).value == "Test question?"


def test_excel_headers_row():
    chunks = [{"text": "content", "attribution": "", "payload": {}}]
    data = chunks_to_excel("Q", chunks)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # Row 1=question, row 2=blank, row 3=headers
    assert ws.cell(row=3, column=1).value == "Source File"
    assert ws.cell(row=3, column=4).value == "Content"


def test_excel_data_row_content():
    chunks = [{
        "text": "Entry 3 yield 87%",
        "attribution": "",
        "payload": {"source_file": "PRJ-031.docx", "page_number": 12, "section": "Table 1"},
    }]
    data = chunks_to_excel("Q", chunks)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # Data starts at row 4
    assert ws.cell(row=4, column=1).value == "PRJ-031.docx"
    assert ws.cell(row=4, column=2).value == 12
    assert ws.cell(row=4, column=3).value == "Table 1"
    assert ws.cell(row=4, column=4).value == "Entry 3 yield 87%"


def test_excel_missing_payload_falls_back_to_attribution():
    chunks = [{"text": "some text", "attribution": "[Source: b.docx]", "payload": {}}]
    data = chunks_to_excel("Q", chunks)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # source_file not in payload -> fall back to attribution string
    assert ws.cell(row=4, column=1).value == "[Source: b.docx]"


def test_excel_multiple_chunks():
    chunks = [
        {"text": f"chunk {i}", "attribution": f"src{i}", "payload": {"source_file": f"f{i}.docx", "page_number": i, "section": f"S{i}"}}
        for i in range(5)
    ]
    data = chunks_to_excel("Q", chunks)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    # 5 data rows starting at row 4
    assert ws.cell(row=8, column=1).value == "f4.docx"


def test_conversation_to_markdown_basic():
    history = [
        {
            "role": "assistant",
            "question": "What catalyst was used?",
            "content": "Pd(OAc)2 was used as the catalyst.",
            "question_type": "factual",
            "chunks": [
                {"attribution": "[Source: PRJ-031.docx | Page 12 | Section: Table 1]"},
            ],
        }
    ]
    result = conversation_to_markdown(history)
    assert isinstance(result, bytes)
    text = result.decode("utf-8")
    assert text.startswith("# Knowledge Base")
    assert "## Q:" in text
    assert "What catalyst was used?" in text
    assert "Pd(OAc)2 was used as the catalyst." in text
    assert "[Source: PRJ-031.docx | Page 12 | Section: Table 1]" in text


def test_conversation_to_markdown_skips_user_messages():
    history = [
        {"role": "user", "content": "What catalyst was used?"},
        {
            "role": "assistant",
            "question": "What catalyst was used?",
            "content": "Pd(OAc)2 was the catalyst.",
            "question_type": "",
            "chunks": [],
        },
    ]
    result = conversation_to_markdown(history)
    text = result.decode("utf-8")
    # Only one Q block should exist
    assert text.count("## Q:") == 1
    # User role should not create its own section header
    assert "## Q: What catalyst was used?" in text


def test_conversation_to_markdown_empty_history():
    result = conversation_to_markdown([])
    assert isinstance(result, bytes)
    assert len(result) > 0
