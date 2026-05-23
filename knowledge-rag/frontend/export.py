# frontend/export.py
"""Export retrieved source chunks to an Excel workbook.

Layout (intended by tests and consumed by the Streamlit sidebar):

    Row 1: ["Question", <question text>, "", ""]
    Row 2: (blank — visual separator)
    Row 3: ["Source File", "Page", "Section", "Content"]  ← header row
    Row 4+: one row per chunk, drawing first from chunk["payload"]
            (where the indexer stores ``source_file``, ``page_number``,
            ``section``) and falling back to chunk["attribution"] /
            chunk["text"] when payload is empty.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADERS = ["Source File", "Page", "Section", "Content"]


def chunks_to_excel(question: str, chunks: list[dict]) -> bytes:
    """Serialise a list of source chunks to an Excel (.xlsx) file.

    Parameters
    ----------
    question : str
        The user question that produced these chunks. Written to row 1.
    chunks : list[dict]
        Chunk dicts with keys ``text``, ``attribution``, ``payload``
        (where ``payload`` may carry ``source_file``, ``page_number``,
        ``section``). All keys are optional; missing values are coerced
        to ``""`` and the attribution string is used as a fallback for
        the source file column.

    Returns
    -------
    bytes
        Raw bytes of the .xlsx file suitable for ``st.download_button``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sources"

    # ── Row 1: Question / value ─────────────────────────────────────────
    ws.cell(row=1, column=1, value="Question").font = Font(bold=True)
    q_cell = ws.cell(row=1, column=2, value=question)
    q_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Row 2 is intentionally blank.

    # ── Row 3: header row ───────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 4+: data rows ───────────────────────────────────────────────
    for i, chunk in enumerate(chunks, start=1):
        row_idx = i + 3  # rows 4, 5, 6, ...
        payload = chunk.get("payload") or {}
        source_file = payload.get("source_file") or chunk.get("attribution") or ""
        page_number = payload.get("page_number", "")
        section = payload.get("section", "")
        text = chunk.get("text", "")

        ws.cell(row=row_idx, column=1, value=source_file)
        ws.cell(row=row_idx, column=2, value=page_number)
        ws.cell(row=row_idx, column=3, value=section)
        text_cell = ws.cell(row=row_idx, column=4, value=text)
        text_cell.alignment = Alignment(wrap_text=True)

        # Alternate row shading for readability.
        if i % 2 == 0:
            fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # ── Column widths ───────────────────────────────────────────────────
    for col, width in enumerate([30, 8, 30, 80], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
